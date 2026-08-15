"""Excel export of student records using openpyxl (hashes only, no plaintext)."""
import base64
import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = [
    "Full Name",
    "Registration ID",
    "Password Hash",
    "Security Question",
    "Security Answer Hash",
    "Department",
    "Year",
    "Semester",
    "Notes Sharing Count",
    "Downloaded Notes Count",
    "Stars Earned",
    "Face Verified",
    "Face Verified At",
    "Verified Face Image",
]


def build_students_workbook(users: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="6D28D9")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for user in users:
        ws.append([
            user.get("fullName", ""),
            user.get("registrationId", ""),
            user.get("passwordHash", ""),
            user.get("securityQuestion", ""),
            user.get("securityAnswerHash", ""),
            user.get("department", ""),
            user.get("year", ""),
            user.get("semester", ""),
            int(user.get("sharedCount", 0)),
            int(user.get("downloadedCount", 0)),
            int(user.get("stars", 0)),
            "YES" if user.get("faceVerified") else "NO",
            user.get("faceVerifiedAt", "") or "",
            "",
        ])
        _attach_face(ws, ws.max_row, len(HEADERS), user.get("faceImage"))

    widths = [26, 18, 60, 34, 60, 16, 12, 12, 20, 24, 14, 16, 26, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _attach_face(ws, row: int, column: int, data_url) -> None:
    """Embed the live-verification capture into the Verified Face Image column."""
    if not data_url or not str(data_url).startswith("data:image/"):
        return
    try:
        raw = base64.b64decode(str(data_url).split(",", 1)[1])
        img = XLImage(io.BytesIO(raw))
        img.width, img.height = 96, 72
        ws.row_dimensions[row].height = 58
        ws.add_image(img, ws.cell(row=row, column=column).coordinate)
    except Exception:  # pragma: no cover - never break the export for a bad image
        ws.cell(row=row, column=column).value = "image unavailable"
